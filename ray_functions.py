"""ray_functions.py, LiftUp projesi için ışın izleme ve simülasyon fonksiyonlarını içerir."""

import numpy as np
import time
import threading
import psutil
import os
from numba import njit, prange
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from mpl_toolkits.mplot3d.art3d import Poly3DCollection
import matplotlib.colors as mcolors
from datetime import datetime


# ------------------------------------------------
# 0. KAYNAK İZLEYİCİ
# ------------------------------------------------
class ResourceMonitor:
    def __init__(self, interval: float = 0.1):
        self.interval    = interval
        self._cpu: list  = []
        self._ram: list  = []
        self._times: list = []
        self._stop_event = threading.Event()
        self._proc       = psutil.Process(os.getpid())
        self._thread     = threading.Thread(target=self._collect, daemon=True)

    def _collect(self):
        t0 = time.perf_counter()
        self._proc.cpu_percent(interval=None)
        self._num_cores = psutil.cpu_count(logical=True) or 1
        while not self._stop_event.is_set():
            cpu_val = self._proc.cpu_percent(interval=None)
            self._cpu.append(cpu_val / self._num_cores)
            mem_mb = self._proc.memory_info().rss / (1024 ** 2)
            self._ram.append(mem_mb)
            self._times.append(time.perf_counter() - t0)
            time.sleep(self.interval)

    def start(self):  self._thread.start()
    def stop(self):
        self._stop_event.set()
        self._thread.join(timeout=2.0)
    def results(self):
        return list(self._cpu), list(self._ram), list(self._times)


# ------------------------------------------------
# 1. UNV DOSYASI OKUMA
# ------------------------------------------------
def read_unv_mesh_optimized(file_path):
    nodes     = {}
    triangles = []

    with open(file_path, 'r') as f:
        iterator = iter(f)
        for line in iterator:
            line = line.strip()
            if line == "2411":
                for line in iterator:
                    line = line.strip()
                    if line == "-1": break
                    node_info = line.split()
                    node_id   = int(node_info[0])
                    coor_line = next(iterator).strip()
                    coor_info = coor_line.split()
                    nodes[node_id] = [float(coor_info[0]),
                                      float(coor_info[1]),
                                      float(coor_info[2])]
            elif line == "2412":
                for line in iterator:
                    line = line.strip()
                    if line == "-1": break
                    point_info   = line.split()
                    point_num    = int(point_info[5])
                    node_ids_line = next(iterator).strip()
                    if point_num == 3:
                        node_ids = [int(x) for x in node_ids_line.split()]
                        triangles.append(node_ids)

    num_triangles = len(triangles)
    A = np.zeros((num_triangles, 3), dtype=np.float64)
    B = np.zeros((num_triangles, 3), dtype=np.float64)
    C = np.zeros((num_triangles, 3), dtype=np.float64)
    for i, t in enumerate(triangles):
        A[i] = nodes[t[0]]
        B[i] = nodes[t[1]]
        C[i] = nodes[t[2]]
    return A, B, C


# ------------------------------------------------
# 2. IŞIN ÜRETME
# ------------------------------------------------
def generate_beam_from_counts(origin, theta_range, phi_range, n_theta, n_phi):
    origin = np.array(origin, dtype=np.float64)
    thetas = np.radians(np.linspace(theta_range[0], theta_range[1], n_theta))
    phis   = np.radians(np.linspace(phi_range[0],   phi_range[1],   n_phi))
    theta_grid, phi_grid = np.meshgrid(thetas, phis, indexing='ij')
    theta_flat = theta_grid.ravel()
    phi_flat   = phi_grid.ravel()
    dx = np.sin(theta_flat) * np.cos(phi_flat)
    dy = np.sin(theta_flat) * np.sin(phi_flat)
    dz = np.cos(theta_flat)
    rays_d = np.stack([dx, dy, dz], axis=1).astype(np.float64)
    rays_o = np.tile(origin, (rays_d.shape[0], 1))
    return rays_o, rays_d


# ------------------------------------------------
# 3. PLÜCKER PRECOMPUTE
# ------------------------------------------------
def precompute_plucker(A, B, C):
    dAB = B - A;  mAB = np.cross(A, dAB)
    dBC = C - B;  mBC = np.cross(B, dBC)
    dCA = A - C;  mCA = np.cross(C, dCA)
    N_arr   = np.cross(dAB, C - A)
    d_plane = np.einsum('ij,ij->i', N_arr, A)
    return (dAB.astype(np.float64), mAB.astype(np.float64),
            dBC.astype(np.float64), mBC.astype(np.float64),
            dCA.astype(np.float64), mCA.astype(np.float64),
            N_arr.astype(np.float64), d_plane.astype(np.float64))


# ------------------------------------------------
# 4. BVH
# ------------------------------------------------
MAX_LEAF_TRIS = 8

def build_bvh(A, B, C):
    num_tris  = A.shape[0]
    centroids = (A + B + C) / 3.0
    node_aabb = []
    node_meta = []

    def compute_aabb(indices):
        pts = np.concatenate([A[indices], B[indices], C[indices]], axis=0)
        return pts.min(axis=0), pts.max(axis=0)

    def build(indices, depth=0):
        node_idx = len(node_aabb)
        mn, mx   = compute_aabb(indices)
        node_aabb.append(np.concatenate([mn, mx]))

        if len(indices) <= MAX_LEAF_TRIS:
            start = len(tri_ids_list)
            tri_ids_list.extend(indices.tolist())
            end   = len(tri_ids_list)
            node_meta.append([-1, end, start])
            return node_idx

        extent = mx - mn
        axis   = int(np.argmax(extent))
        order  = np.argsort(centroids[indices, axis])
        sorted_indices = indices[order]
        mid    = len(sorted_indices) // 2

        node_meta.append([-1, -1, -1])
        left_child  = build(sorted_indices[:mid],  depth + 1)
        right_child = build(sorted_indices[mid:],  depth + 1)
        node_meta[node_idx] = [left_child, right_child, -1]
        return node_idx

    tri_ids_list = []
    build(np.arange(num_tris, dtype=np.int32))

    return (np.array(node_aabb, dtype=np.float64),
            np.array(node_meta, dtype=np.int32),
            np.array(tri_ids_list, dtype=np.int32))


# ------------------------------------------------
# 5. IŞIN-KUTU KESİŞİM
# ------------------------------------------------
@njit(fastmath=True)
def ray_aabb_intersect(ox, oy, oz, idx, idy, idz, mn, mx, t_max):
    tx1 = (mn[0]-ox)*idx;  tx2 = (mx[0]-ox)*idx
    t_near = min(tx1,tx2); t_far = max(tx1,tx2)
    ty1 = (mn[1]-oy)*idy;  ty2 = (mx[1]-oy)*idy
    t_near = max(t_near,min(ty1,ty2)); t_far = min(t_far,max(ty1,ty2))
    tz1 = (mn[2]-oz)*idz;  tz2 = (mx[2]-oz)*idz
    t_near = max(t_near,min(tz1,tz2)); t_far = min(t_far,max(tz1,tz2))
    return t_far >= max(t_near, 1e-9) and t_near < t_max


# ------------------------------------------------
# 6. PLÜCKER + BVH TRACER
# ------------------------------------------------
@njit(parallel=True, fastmath=True)
def plucker_bvh_tracer(rays_o, rays_d,
                        dAB, mAB, dBC, mBC, dCA, mCA,
                        N_arr, d_plane,
                        bvh_nodes, bvh_meta, bvh_tids):
    num_rays = rays_o.shape[0]
    hit_triangle_ids = np.full(num_rays, -1,     dtype=np.int32)
    hit_distances    = np.full(num_rays, np.inf, dtype=np.float64)

    for i in prange(num_rays):
        ox = rays_o[i,0]; oy = rays_o[i,1]; oz = rays_o[i,2]
        dx = rays_d[i,0]; dy = rays_d[i,1]; dz = rays_d[i,2]
        mRx = oy*dz - oz*dy; mRy = oz*dx - ox*dz; mRz = ox*dy - oy*dx
        idx = 1.0/dx if abs(dx)>1e-12 else 1e12
        idy = 1.0/dy if abs(dy)>1e-12 else 1e12
        idz = 1.0/dz if abs(dz)>1e-12 else 1e12

        closest_t   = np.inf
        closest_tri = -1
        stack    = np.empty(64, dtype=np.int32)
        stack[0] = 0; sp = 1

        while sp > 0:
            sp -= 1
            node_idx = stack[sp]
            mn = bvh_nodes[node_idx,:3]; mx = bvh_nodes[node_idx,3:]
            if not ray_aabb_intersect(ox,oy,oz,idx,idy,idz,mn,mx,closest_t):
                continue
            left_child = bvh_meta[node_idx,0]
            tri_end    = bvh_meta[node_idx,1]
            tri_start  = bvh_meta[node_idx,2]

            if left_child == -1:
                for k in range(tri_start, tri_end):
                    j = bvh_tids[k]
                    s0 = (dx*mAB[j,0]+dy*mAB[j,1]+dz*mAB[j,2] +
                          dAB[j,0]*mRx+dAB[j,1]*mRy+dAB[j,2]*mRz)
                    s1 = (dx*mBC[j,0]+dy*mBC[j,1]+dz*mBC[j,2] +
                          dBC[j,0]*mRx+dBC[j,1]*mRy+dBC[j,2]*mRz)
                    if (s0>0. and s1<0.) or (s0<0. and s1>0.): continue
                    s2 = (dx*mCA[j,0]+dy*mCA[j,1]+dz*mCA[j,2] +
                          dCA[j,0]*mRx+dCA[j,1]*mRy+dCA[j,2]*mRz)
                    if not ((s0>=0. and s1>=0. and s2>=0.) or
                            (s0<=0. and s1<=0. and s2<=0.)): continue
                    Nx=N_arr[j,0]; Ny=N_arr[j,1]; Nz=N_arr[j,2]
                    denom = dx*Nx+dy*Ny+dz*Nz
                    if abs(denom) < 1e-9: continue
                    t = (d_plane[j]-(ox*Nx+oy*Ny+oz*Nz)) / denom
                    if t < 1e-9 or t >= closest_t: continue
                    closest_t=t; closest_tri=j
            else:
                stack[sp]=left_child; stack[sp+1]=tri_end; sp+=2

        if closest_tri != -1:
            hit_triangle_ids[i] = closest_tri
            hit_distances[i]    = closest_t

    return hit_triangle_ids, hit_distances


# ------------------------------------------------
# 7. MULTI-BOUNCE SİMÜLASYONU
# ------------------------------------------------
def run_simulation(rays_o, rays_d, A, B, C,
                   dAB, mAB, dBC, mBC, dCA, mCA,
                   N_arr, d_plane,
                   bvh_nodes, bvh_meta, bvh_tids,
                   max_safety_limit=100):

    num_triangles = A.shape[0]
    hit_counts  = np.zeros(num_triangles, dtype=np.int32)
    angle_sum   = np.zeros(num_triangles, dtype=np.float64)
    angle_count = np.zeros(num_triangles, dtype=np.int32)
    bounce_paths = []
    
    # Gelen ve seken açı verileri
    incoming_angles = []    # Her çarpışmadaki gelen açı
    reflection_angles = []  # Her çarpışmadaki seken açı
    reflection_tri_ids = [] # Seken açının hangi üçgene ait olduğu
    
    # İlk çarpışma bilgisini sakla (renklendirme için)
    first_hit_ids = None
    initial_rays_o = rays_o.copy()
    initial_rays_d = rays_d.copy()
    
    # Sekme bazlı zamanlama ve hit oranı
    bounce_times = []  # Her sekmenin süresi
    bounce_hit_rates = []  # Her sekmenin hit oranı

    print(f"\nIşınlar kaviteden çıkana kadar serbest bırakıldı!")
    current_rays_o = rays_o.copy()
    current_rays_d = rays_d.copy()

    monitor = ResourceMonitor(interval=0.1)
    monitor.start()
    total_sim_start = time.perf_counter()

    bounce = 0
    while bounce < max_safety_limit:
        bounce += 1
        num_active = current_rays_o.shape[0]
        print(f"  --> Sekme {bounce}: Kavite içinde {num_active} ışın...")

        # Sekme zamanlaması
        bounce_start = time.perf_counter()
        
        hit_ids, hit_dists = plucker_bvh_tracer(
            current_rays_o, current_rays_d,
            dAB, mAB, dBC, mBC, dCA, mCA,
            N_arr, d_plane, bvh_nodes, bvh_meta, bvh_tids
        )
        
        bounce_time = time.perf_counter() - bounce_start
        bounce_times.append(bounce_time)
        
        # İlk sekmede hangi ışınların çarptığını kaydet
        if bounce == 1:
            first_hit_ids = hit_ids.copy()

        valid_hits = hit_ids != -1
        num_hits = int(np.sum(valid_hits))
        hit_rate = (num_hits / num_active * 100.0) if num_active > 0 else 0.0
        bounce_hit_rates.append(hit_rate)
        
        print(f"      Süre: {bounce_time:.4f} sn  |  Hit Oranı: %{hit_rate:.1f}")
        
        if num_hits == 0:
            print(f"  --> SONUÇ: Tüm ışınlar kaviteden çıktı! ({bounce-1}. sekmede bitti)")
            break

        valid_hit_ids = hit_ids[valid_hits]
        valid_rays_o  = current_rays_o[valid_hits]
        valid_rays_d  = current_rays_d[valid_hits]
        valid_dists   = hit_dists[valid_hits]
        hit_points    = valid_rays_o + valid_rays_d * valid_dists[:,np.newaxis]

        normals      = N_arr[valid_hit_ids].copy()
        norm_lengths = np.linalg.norm(normals, axis=1)
        norm_lengths[norm_lengths==0] = 1.0
        normals /= norm_lengths[:,np.newaxis]

        I           = valid_rays_d
        dot_product = np.einsum('ij,ij->i', I, normals)
        cos_angles  = np.clip(np.abs(dot_product), 0.0, 1.0)
        angles_deg  = np.degrees(np.arccos(cos_angles))
        
        # Gelen açıları kaydet
        incoming_angles.extend(angles_deg.tolist())

        # Yansıyan ışın yönünü hesapla
        R = I - 2.0 * dot_product[:,np.newaxis] * normals
        
        # Seken açıyı hesapla (yansıyan ışın ile normal arasındaki açı)
        dot_product_reflect = np.einsum('ij,ij->i', R, normals)
        cos_angles_reflect  = np.clip(np.abs(dot_product_reflect), 0.0, 1.0)
        reflect_angles_deg  = np.degrees(np.arccos(cos_angles_reflect))
        
        # Seken açıları kaydet (üçgen ID'siyle birlikte)
        reflection_angles.extend(reflect_angles_deg.tolist())
        reflection_tri_ids.extend(valid_hit_ids.tolist())

        for k, hit_id in enumerate(valid_hit_ids):
            hit_counts[hit_id]  += 1
            angle_sum[hit_id]   += angles_deg[k]
            angle_count[hit_id] += 1

        bounce_paths.append((valid_rays_o, hit_points))

        current_rays_o = hit_points + R * 1e-4
        current_rays_d = R

    total_sim_end = time.perf_counter()
    monitor.stop()
    cpu_samples, ram_samples, time_samples = monitor.results()
    total_elapsed = total_sim_end - total_sim_start

    print(f"\n--- SİMÜLASYON BİTTİ ---")
    print(f"Toplam Süre: {total_elapsed:.4f} sn")
    if cpu_samples:
        print(f"CPU  Ort: {np.mean(cpu_samples):.1f}%  Maks: {np.max(cpu_samples):.1f}%")
        print(f"RAM  Ort: {np.mean(ram_samples):.0f} MB  Maks: {np.max(ram_samples):.0f} MB")
    
    # Sekme istatistikleri
    if bounce_times:
        print(f"\n--- SEKME İSTATİSTİKLERİ ---")
        print(f"Sekme Sayısı            : {len(bounce_times)}")
        print(f"Ortalama Sekme Süresi   : {np.mean(bounce_times):.4f} sn")
        print(f"Min  Sekme Süresi       : {np.min(bounce_times):.4f} sn")
        print(f"Max  Sekme Süresi       : {np.max(bounce_times):.4f} sn")
        print(f"Std  Sekme Süresi       : {np.std(bounce_times):.4f} sn")
        print(f"Ortalama Hit Oranı      : %{np.mean(bounce_hit_rates):.1f}")
        print(f"Son Sekme Hit Oranı     : %{bounce_hit_rates[-1]:.1f}")

    avg_angles = np.zeros(num_triangles, dtype=np.float64)
    mask = angle_count > 0
    avg_angles[mask] = angle_sum[mask] / angle_count[mask]

    resource_data = {
        "cpu":  cpu_samples, "ram":   ram_samples,
        "times":time_samples,"duration":total_elapsed,
        "incoming_angles":    np.array(incoming_angles),
        "reflection_angles":  np.array(reflection_angles),
        "reflection_tri_ids": np.array(reflection_tri_ids, dtype=np.int32),
        "bounce_times":      np.array(bounce_times),
        "bounce_hit_rates":  np.array(bounce_hit_rates),
    }
    return hit_counts, avg_angles, bounce_paths, resource_data, first_hit_ids, initial_rays_o, initial_rays_d


# ------------------------------------------------
# 8. GÖRSELLEŞTİRME
# ------------------------------------------------
def visualize(A, B, C, hit_counts, avg_angles, bounce_paths,
              origin, theta_range, phi_range, n_theta, n_phi,
              min_x, max_x, min_y, max_y, min_z, max_z,
              resource_data=None, cam_elev=20, cam_azim=45):

    total_rays  = n_theta * n_phi
    active_hits = hit_counts > 0

    if np.sum(active_hits) > 0:
        print(f"\n--- AÇI İSTATİSTİKLERİ ---")
        print(f"  Ortalama : {np.mean(avg_angles[active_hits]):.2f}°")
        print(f"  Min      : {np.min(avg_angles[active_hits]):.2f}°")
        print(f"  Max      : {np.max(avg_angles[active_hits]):.2f}°")
        print(f"  Çarpılan : {np.sum(active_hits)} üçgen")

    has_resource = (resource_data is not None
                    and len(resource_data.get("cpu",[])) > 1)

    if has_resource:
        fig = plt.figure(figsize=(22,14))
        gs  = gridspec.GridSpec(2, 3, figure=fig,
                                height_ratios=[1.15, 0.85],
                                hspace=0.38, wspace=0.35)
        fig.subplots_adjust(left=0.05,right=0.97,top=0.92,bottom=0.06)
    else:
        fig = plt.figure(figsize=(22,9))
        gs  = gridspec.GridSpec(1, 2, width_ratios=[1.1,1], figure=fig)
        fig.subplots_adjust(left=0.05,right=0.97,top=0.90,bottom=0.08,wspace=0.35)

    fig.suptitle(
        f"PLÜCKER + PRECOMPUTE + BVH  |  θ: {theta_range}°  |  φ: {phi_range}°  |  "
        f"Toplam Işın: {n_theta}×{n_phi} = {total_rays}",
        fontsize=12, fontweight='bold'
    )

    ax1 = fig.add_subplot(gs[0,0] if has_resource else gs[0], projection='3d')
    ax1.set_title('Vuruş Sayısı Isı Haritası', fontweight='bold')

    if np.sum(active_hits) > 0:
        hit_A  = A[active_hits]; hit_B = B[active_hits]; hit_C = C[active_hits]
        counts = hit_counts[active_hits]
        verts  = [[hit_A[i], hit_B[i], hit_C[i]] for i in range(len(hit_A))]
        norm1  = mcolors.Normalize(vmin=1, vmax=np.max(counts))
        cmap1  = plt.get_cmap('jet')
        ax1.add_collection3d(Poly3DCollection(verts, facecolors=cmap1(norm1(counts)),
                                               edgecolors='none', alpha=0.9))
        mappable1 = plt.cm.ScalarMappable(norm=norm1, cmap=cmap1)
        mappable1.set_array(counts)
        fig.colorbar(mappable1, ax=ax1, shrink=0.5, aspect=15,
                     label='Vuruş Sayısı')

    for start_pts, end_pts in bounce_paths:
        n    = start_pts.shape[0]
        step = max(1, n // 100)
        for i in range(0, n, step):
            ax1.plot([start_pts[i,0],end_pts[i,0]],
                     [start_pts[i,1],end_pts[i,1]],
                     [start_pts[i,2],end_pts[i,2]],
                     color='magenta', alpha=0.4, linewidth=0.5)

    ax1.scatter(*origin, color='cyan', s=300, marker='X',
                label='Radar Kaynağı', depthshade=False)
    ax1.scatter(A[::10,0],A[::10,1],A[::10,2],
                color='gray', s=0.1, alpha=0.05)
    ax1.set_xlabel('X (m)'); ax1.set_ylabel('Y (m)'); ax1.set_zlabel('Z (m)')
    ax1.set_xlim([min_x,max_x]); ax1.set_ylim([min_y,max_y]); ax1.set_zlim([min_z,max_z])
    ax1.view_init(elev=cam_elev, azim=cam_azim); ax1.legend(fontsize=8)

    ax2 = fig.add_subplot(gs[0,1] if has_resource else gs[1])
    ax2.set_title('Sekme Açıları Dağılımı\n(0°=Dik Çarpış  →  90°=Sıyırma)',
                  fontweight='bold')
    if np.sum(active_hits) > 0:
        angles = avg_angles[active_hits]
        ax2.hist(angles, bins=45, range=(0,90),
                 color='steelblue', edgecolor='white', linewidth=0.4, alpha=0.85)
        ax2.axvline(np.mean(angles), color='red', linestyle='--', linewidth=1.5,
                    label=f'Ort: {np.mean(angles):.1f}°')
        ax2.axvline(np.median(angles), color='orange', linestyle=':', linewidth=1.5,
                    label=f'Medyan: {np.median(angles):.1f}°')
        ax2.set_xlabel('Gelme Açısı (°)', fontsize=11)
        ax2.set_ylabel('Üçgen Sayısı', fontsize=11)
        ax2.set_xlim(0,90); ax2.legend(fontsize=9)
        ax2.grid(axis='y', linestyle='--', alpha=0.4)

    if has_resource:
        import psutil as _ps
        cpu_arr = np.array(resource_data["cpu"])
        ram_arr = np.array(resource_data["ram"])
        t_arr   = np.array(resource_data["times"])
        ram_gb  = ram_arr / 1024.0
        total_ram_gb = _ps.virtual_memory().total / (1024**3)

        ax3 = fig.add_subplot(gs[1,0])
        ax3.set_title('CPU Kullanımı — Zaman Serisi', fontweight='bold')
        ax3.plot(t_arr, cpu_arr, color='#2196F3', linewidth=1.0, alpha=0.8)
        ax3.fill_between(t_arr, cpu_arr, alpha=0.15, color='#2196F3')
        ax3.axhline(np.mean(cpu_arr),color='red',linestyle='--',linewidth=1.2,
                    label=f'Ort: {np.mean(cpu_arr):.1f}%')
        ax3.axhline(np.max(cpu_arr),color='orange',linestyle=':',linewidth=1.0,
                    label=f'Maks: {np.max(cpu_arr):.1f}%')
        ax3.set_xlabel('Süre (sn)'); ax3.set_ylabel('CPU (%)')
        ax3.set_ylim(0,max(100,np.max(cpu_arr)*1.15))
        ax3.set_xlim(0,t_arr[-1]); ax3.legend(fontsize=8)
        ax3.grid(linestyle='--',alpha=0.35)

        ax4 = fig.add_subplot(gs[1,1])
        ax4.set_title('RAM Kullanımı — Zaman Serisi', fontweight='bold')
        ax4.plot(t_arr, ram_gb, color='#4CAF50', linewidth=1.0, alpha=0.8)
        ax4.fill_between(t_arr, ram_gb, alpha=0.15, color='#4CAF50')
        ax4.axhline(np.mean(ram_gb),color='red',linestyle='--',linewidth=1.2,
                    label=f'Ort: {np.mean(ram_gb):.2f} GB')
        ax4.axhline(np.max(ram_gb),color='orange',linestyle=':',linewidth=1.0,
                    label=f'Maks: {np.max(ram_gb):.2f} GB')
        ax4.set_xlabel('Süre (sn)'); ax4.set_ylabel('RAM (GB)')
        ax4.set_xlim(0,t_arr[-1]); ax4.legend(fontsize=8)
        ax4.grid(linestyle='--',alpha=0.35)

        ax5 = fig.add_subplot(gs[1,2])
        ax5.axis('off')
        ax5.set_title('Kaynak Özeti', fontweight='bold')
        ram_delta = np.max(ram_gb) - ram_gb[0]
        pct_used  = (np.max(ram_gb) / total_ram_gb) * 100.0
        lines = [
            ("── CPU ──",""),("Ortalama",f"{np.mean(cpu_arr):.1f}%"),
            ("Maksimum",f"{np.max(cpu_arr):.1f}%"),
            ("Minimum",f"{np.min(cpu_arr):.1f}%"),
            ("",""),("── RAM ──",""),
            ("Ortalama",f"{np.mean(ram_gb):.2f} GB"),
            ("Maksimum",f"{np.max(ram_gb):.2f} GB"),
            ("Simülasyon artışı",f"+{ram_delta:.2f} GB"),
            ("Toplam RAM",f"{total_ram_gb:.1f} GB  ({pct_used:.1f}%)"),
            ("",""),("── GENEL ──",""),
            ("Simülasyon süresi",f"{resource_data['duration']:.2f} sn"),
            ("Örnekler",f"{len(cpu_arr)}"),
        ]
        y = 0.97
        for label, value in lines:
            if label.startswith("──"):
                ax5.text(0.05,y,label,transform=ax5.transAxes,
                         fontsize=9,fontweight='bold',color='#333333',va='top')
            elif label:
                ax5.text(0.05,y,label,transform=ax5.transAxes,
                         fontsize=9,color='#555555',va='top')
                ax5.text(0.65,y,value,transform=ax5.transAxes,
                         fontsize=9,fontweight='bold',color='#111111',va='top',ha='left')
            y -= 0.068

    plt.show()


# ------------------------------------------------
# 8B. YENİ GÖRSELLEŞTİRME FONKSİYONLARI (GUI İÇİN)
# ------------------------------------------------

def create_ray_path_figure(A, B, C, bounce_paths, origin,
                           rays_o=None, rays_d=None, first_hit_ids=None,
                           cam_elev=20, cam_azim=45,
                           bg_color="#FFFFFF", text_color="#333333"):
    """
    Işın yollarını gösteren 3D figure oluştur.
    Üçgenler transparan, ışın yolları renkli çizgiler olarak gösterilir.
    Eğer verilirse, çarpan ışınlar kırmızı, çarpmayan ışınlar mavi gösterilir.
    """
    fig = plt.Figure(figsize=(10, 8), facecolor=bg_color, dpi=80)
    ax = fig.add_subplot(111, projection='3d')
    ax.set_facecolor(bg_color)
    
    # Mesh sınırlarını hesapla
    all_pts = np.concatenate([A, B, C], axis=0)
    mn, mx = all_pts.min(0), all_pts.max(0)
    mesh_center = (mn + mx) / 2.0
    mesh_scale = float((mx - mn).max())
    mr = mesh_scale / 2.0
    
    # Üçgenleri transparan gösterir (ÇOK DAHA AZ NOKTA)
    step = max(1, len(A) // 1000)  # 3000'den 1000'e düşürüldü
    pts = np.concatenate([A[::step], B[::step], C[::step]], axis=0)
    ax.scatter(pts[:, 0], pts[:, 1], pts[:, 2],
               c='#5A7399', s=0.2, alpha=0.1, linewidths=0, rasterized=True)
    
    # Işın yollarını çiz - Renklendirme mantığı
    use_color_coding = (rays_o is not None and rays_d is not None and first_hit_ids is not None)
    
    if use_color_coding:
        # YENİ DAVRANŞ: Çarpan/çarpmayan ışınları farklı renklerde göster
        
        # 1. ÇARPMAYAN IŞINLARI MAVİ ÇİZ (önce, altta kalsın)
        miss_ray_indices = np.where(first_hit_ids == -1)[0]
        max_miss_rays = 50  # Performans için limit
        
        if len(miss_ray_indices) > 0:
            step = max(1, len(miss_ray_indices) // max_miss_rays)
            for idx in miss_ray_indices[::step]:
                start = rays_o[idx]
                direction = rays_d[idx]
                # Mesh dışına uzanan bir çizgi çiz
                end = start + direction * mesh_scale * 0.8
                ax.plot([start[0], end[0]],
                       [start[1], end[1]],
                       [start[2], end[2]],
                       color='#4D9EFF', alpha=0.4, linewidth=0.5,
                       rasterized=True)
        
        # 2. ÇARPAN IŞINLARI KIRMIZI ÇİZ (bounce_paths kullanarak, üstte)
        max_hit_rays = 80
        for bounce_idx, (start_pts, end_pts) in enumerate(bounce_paths):
            n = start_pts.shape[0]
            step = max(1, n // max_hit_rays)
            for i in range(0, min(n, max_hit_rays * step), step):
                ax.plot([start_pts[i,0], end_pts[i,0]],
                       [start_pts[i,1], end_pts[i,1]],
                       [start_pts[i,2], end_pts[i,2]],
                       color='#E74C3C', alpha=0.6, linewidth=0.7,
                       rasterized=True)
    else:
        # Tüm ışınları mavi göster
        max_rays_to_draw = 80
        for bounce_idx, (start_pts, end_pts) in enumerate(bounce_paths):
            n = start_pts.shape[0]
            step = max(1, n // max_rays_to_draw)
            for i in range(0, min(n, max_rays_to_draw * step), step):
                ax.plot([start_pts[i,0], end_pts[i,0]],
                       [start_pts[i,1], end_pts[i,1]],
                       [start_pts[i,2], end_pts[i,2]],
                       color='#4D9EFF', alpha=0.5, linewidth=0.6, rasterized=True)
    
    # Kaynak noktayı vurgula
    ax.scatter(*origin, color='#2ECC71', s=200, marker='X',
               label='Işın Kaynağı', depthshade=False, edgecolors='white', linewidths=1)
    
    # Eksen ayarları
    ax.set_xlabel('X (m)', color=text_color, fontsize=9)
    ax.set_ylabel('Y (m)', color=text_color, fontsize=9)
    ax.set_zlabel('Z (m)', color=text_color, fontsize=9)
    ax.set_xlim(mesh_center[0]-mr, mesh_center[0]+mr)
    ax.set_ylim(mesh_center[1]-mr, mesh_center[1]+mr)
    ax.set_zlim(mesh_center[2]-mr, mesh_center[2]+mr)
    ax.view_init(elev=cam_elev, azim=cam_azim)
    
    # Grid ve stil
    ax.grid(True, color='#D0D0D0', linewidth=0.3, alpha=0.3)
    for axis in [ax.xaxis, ax.yaxis, ax.zaxis]:
        axis.pane.fill = False
        axis.pane.set_edgecolor('#D0D0D0')
        axis.line.set_color('#B0B0B0')
    ax.tick_params(colors='#666666', labelsize=8)
    
    # Legend - Renklendirme varsa özel legend, yoksa basit
    if use_color_coding:
        from matplotlib.lines import Line2D
        legend_elements = [
            Line2D([0], [0], color='#E74C3C', linewidth=2, label='Kaviteye Giren'),
            Line2D([0], [0], color='#4D9EFF', linewidth=2, label='Girmeyen'),
            Line2D([0], [0], marker='X', color='w', markerfacecolor='#2ECC71',
                   markersize=10, label='Işın Kaynağı', linestyle='None')
        ]
        ax.legend(handles=legend_elements, fontsize=9, facecolor='#F5F5F5',
                 labelcolor=text_color, edgecolor='#D0D0D0', loc='upper right')
    else:
        ax.legend(fontsize=9, facecolor='#F5F5F5', labelcolor=text_color,
                 edgecolor='#D0D0D0', loc='upper right')
    
    try: ax.dist = 8
    except: pass
    
    fig.suptitle('Işın Yolu Görünümü', color=text_color, fontsize=13, fontweight='bold')
    fig.tight_layout()
    
    return fig


def create_heatmap_figure(A, B, C, hit_counts,
                         cam_elev=20, cam_azim=45,
                         bg_color="#FFFFFF", text_color="#333333"):
    """
    Isı haritası figure oluştur — geliştirilmiş versiyon.
    - inferno renk haritası (algısal olarak eşit)
    - Logaritmik normalizasyon (fark büyükse)
    - İnce kenar çizgileri (3D derinlik)
    - Pasif mesh yarı saydam yüzey olarak
    - Büyük ve okunaklı colorbar
    - Max çarpış noktası vurgusu
    - Başlıkta istatistik özeti
    """
    fig = plt.Figure(figsize=(12, 9), facecolor=bg_color, dpi=100)
    ax = fig.add_subplot(111, projection='3d')
    ax.set_facecolor(bg_color)

    # Mesh sınırlarını hesapla
    all_pts = np.concatenate([A, B, C], axis=0)
    mn, mx = all_pts.min(0), all_pts.max(0)
    mesh_center = (mn + mx) / 2.0
    mesh_scale = float((mx - mn).max())
    mr = mesh_scale / 2.0

    # Çarpılan üçgenleri bul
    active_hits = hit_counts > 0

    if np.sum(active_hits) > 0:
        hit_A = A[active_hits]
        hit_B = B[active_hits]
        hit_C = C[active_hits]
        counts = hit_counts[active_hits]

        # Çok fazla üçgen varsa orantılı örnekle (üst sınır yükseltildi)
        max_triangles = 15000
        if len(hit_A) > max_triangles:
            step = len(hit_A) // max_triangles
            hit_A = hit_A[::step]
            hit_B = hit_B[::step]
            hit_C = hit_C[::step]
            counts = counts[::step]

        verts = [[hit_A[i], hit_B[i], hit_C[i]] for i in range(len(hit_A))]

        # inferno: algısal olarak eşit, koyu→açık geçiş, renk körü dostu
        cmap = plt.get_cmap('inferno')

        # Logaritmik normalizasyon: aralık 10x'den büyükse log kullan
        if np.max(counts) / max(np.min(counts), 1) > 10:
            norm = mcolors.LogNorm(vmin=1, vmax=np.max(counts))
        else:
            norm = mcolors.Normalize(vmin=1, vmax=np.max(counts))

        ax.add_collection3d(Poly3DCollection(
            verts,
            facecolors=cmap(norm(counts)),
            edgecolors='#00000022',   # ince yarı saydam kenar → 3D derinlik
            linewidths=0.2,
            alpha=0.92,
            rasterized=True
        ))

        # Colorbar — daha büyük ve okunaklı
        mappable = plt.cm.ScalarMappable(norm=norm, cmap=cmap)
        mappable.set_array(counts)
        cbar = fig.colorbar(mappable, ax=ax, shrink=0.75, aspect=15, pad=0.05)
        cbar.set_label('Çarpma Sayısı', color=text_color, fontsize=11)
        cbar.ax.tick_params(colors=text_color, labelsize=9)
        cbar.outline.set_edgecolor('#A0A0A0')

        # En fazla çarpılan üçgeni beyaz nokta + kırmızı kenarla vurgula
        max_idx = np.argmax(counts)
        cx = (hit_A[max_idx][0] + hit_B[max_idx][0] + hit_C[max_idx][0]) / 3
        cy = (hit_A[max_idx][1] + hit_B[max_idx][1] + hit_C[max_idx][1]) / 3
        cz = (hit_A[max_idx][2] + hit_B[max_idx][2] + hit_C[max_idx][2]) / 3
        ax.scatter([cx], [cy], [cz], color='white', s=60, zorder=10,
                   edgecolors='red', linewidths=1.5,
                   label=f'Max: {int(np.max(counts))} çarpış')
        ax.legend(loc='upper left', fontsize=9,
                  facecolor=bg_color, labelcolor=text_color)

    # Çarpılmayan üçgenleri yarı saydam yüzey olarak göster (scatter yerine)
    inactive = ~active_hits
    if np.sum(inactive) > 0:
        step = max(1, np.sum(inactive) // 3000)
        in_A = A[inactive][::step]
        in_B = B[inactive][::step]
        in_C = C[inactive][::step]
        inactive_verts = [[in_A[i], in_B[i], in_C[i]] for i in range(len(in_A))]
        ax.add_collection3d(Poly3DCollection(
            inactive_verts,
            facecolors='#CCCCCC',
            edgecolors='none',
            alpha=0.12,
            rasterized=True
        ))

    # Eksen ayarları
    ax.set_xlabel('X (m)', color=text_color, fontsize=10)
    ax.set_ylabel('Y (m)', color=text_color, fontsize=10)
    ax.set_zlabel('Z (m)', color=text_color, fontsize=10)
    ax.set_xlim(mesh_center[0]-mr, mesh_center[0]+mr)
    ax.set_ylim(mesh_center[1]-mr, mesh_center[1]+mr)
    ax.set_zlim(mesh_center[2]-mr, mesh_center[2]+mr)
    ax.view_init(elev=cam_elev, azim=cam_azim)

    # Grid ve stil
    ax.grid(True, color='#D0D0D0', linewidth=0.3, alpha=0.3)
    for axis in [ax.xaxis, ax.yaxis, ax.zaxis]:
        axis.pane.fill = False
        axis.pane.set_edgecolor('#D0D0D0')
        axis.line.set_color('#B0B0B0')
    ax.tick_params(colors='#666666', labelsize=8)

    # Kamera mesafesi mesh büyüklüğüne göre otomatik ayarla
    try:
        ax.dist = max(7, min(11, 9 - mesh_scale * 0.5))
    except Exception:
        pass

    # Başlıkta istatistik özeti
    n_hit = int(np.sum(active_hits))
    n_total = len(hit_counts)
    hit_pct = 100 * n_hit / max(n_total, 1)
    fig.suptitle(
        f'Isı Haritası  |  {n_hit:,} / {n_total:,} üçgen çarpıldı  ({hit_pct:.1f}%)',
        color=text_color, fontsize=13, fontweight='bold'
    )
    fig.tight_layout()

    return fig


def create_ray_trace_figure(A, B, C, bounce_paths, hit_counts, origin,
                           cam_elev=20, cam_azim=45,
                           bg_color="#FFFFFF", text_color="#333333"):
    """Işın izi figure oluştur.
    Işınların yönünü ve yoğunluğunu birlikte gösterir."""
    fig = plt.Figure(figsize=(10, 8), facecolor=bg_color, dpi=80)
    ax = fig.add_subplot(111, projection='3d')
    ax.set_facecolor(bg_color)
    
    # Mesh sınırlarını hesapla
    all_pts = np.concatenate([A, B, C], axis=0)
    mn, mx = all_pts.min(0), all_pts.max(0)
    mesh_center = (mn + mx) / 2.0
    mesh_scale = float((mx - mn).max())
    mr = mesh_scale / 2.0
    
    # Çarpılan üçgenleri hafif renkli göster (PERFORMANS: Daha az üçgen)
    active_hits = hit_counts > 0
    if np.sum(active_hits) > 0:
        hit_A = A[active_hits]
        hit_B = B[active_hits]
        hit_C = C[active_hits]
        counts = hit_counts[active_hits]
        
        # PERFORMANS: Çok fazla üçgen varsa downsample yap
        max_triangles = 3000
        if len(hit_A) > max_triangles:
            indices = np.random.choice(len(hit_A), max_triangles, replace=False)
            hit_A = hit_A[indices]
            hit_B = hit_B[indices]
            hit_C = hit_C[indices]
            counts = counts[indices]
        
        verts = [[hit_A[i], hit_B[i], hit_C[i]] for i in range(len(hit_A))]
        norm = mcolors.Normalize(vmin=1, vmax=np.max(counts))
        cmap = plt.get_cmap('viridis')
        
        ax.add_collection3d(Poly3DCollection(verts,
                                            facecolors=cmap(norm(counts)),
                                            edgecolors='none',
                                            alpha=0.3,
                                            rasterized=True))
    
    # Işın yollarını yoğunluk ile göster (ÇOK DAHA AZ IŞIN)
    max_rays_per_bounce = 60  # Her sekme için maksimum ışın
    for bounce_idx, (start_pts, end_pts) in enumerate(bounce_paths):
        n = start_pts.shape[0]
        step = max(1, n // max_rays_per_bounce)
        
        # Her sekme için farklı renk tonu
        alpha = max(0.3, 1.0 - bounce_idx * 0.1)
        color_intensity = max(0.4, 1.0 - bounce_idx * 0.15)
        
        for i in range(0, min(n, max_rays_per_bounce * step), step):
            # Işın yönünü ok ile göster
            direction = end_pts[i] - start_pts[i]
            length = np.linalg.norm(direction)
            if length > 1e-6:
                direction = direction / length
                # Çizgi
                ax.plot([start_pts[i,0], end_pts[i,0]],
                       [start_pts[i,1], end_pts[i,1]],
                       [start_pts[i,2], end_pts[i,2]],
                       color=plt.cm.plasma(color_intensity),
                       alpha=alpha, linewidth=0.5, rasterized=True)
    
    # Kaynak noktayı vurgula
    ax.scatter(*origin, color='#2ECC71', s=200, marker='X',
               label='Işın Kaynağı', depthshade=False, edgecolors='white', linewidths=1)
    
    # Eksen ayarları
    ax.set_xlabel('X (m)', color=text_color, fontsize=9)
    ax.set_ylabel('Y (m)', color=text_color, fontsize=9)
    ax.set_zlabel('Z (m)', color=text_color, fontsize=9)
    ax.set_xlim(mesh_center[0]-mr, mesh_center[0]+mr)
    ax.set_ylim(mesh_center[1]-mr, mesh_center[1]+mr)
    ax.set_zlim(mesh_center[2]-mr, mesh_center[2]+mr)
    ax.view_init(elev=cam_elev, azim=cam_azim)
    
    # Grid ve stil
    ax.grid(True, color='#D0D0D0', linewidth=0.3, alpha=0.3)
    for axis in [ax.xaxis, ax.yaxis, ax.zaxis]:
        axis.pane.fill = False
        axis.pane.set_edgecolor('#D0D0D0')
        axis.line.set_color('#B0B0B0')
    ax.tick_params(colors='#666666', labelsize=8)
    ax.legend(fontsize=9, facecolor='#F5F5F5', labelcolor=text_color,
             edgecolor='#D0D0D0', loc='upper right')
    
    try: ax.dist = 8
    except: pass
    
    fig.suptitle('Işın İzi Görünümü', color=text_color, fontsize=13, fontweight='bold')
    fig.tight_layout()
    
    return fig


def create_angle_histogram(avg_angles, hit_counts,
                           bg_color="#FFFFFF", text_color="#333333"):
    """Gelme açısı histogramı oluştur.
    0° = Dik çarpış, 90° = Sıyırma"""
    fig = plt.Figure(figsize=(10, 6), facecolor=bg_color, dpi=80)
    ax = fig.add_subplot(111)
    ax.set_facecolor('#FFFFFF')
    
    active_hits = hit_counts > 0
    
    if np.sum(active_hits) > 0:
        angles = avg_angles[active_hits]
        
        # Histogram
        n, bins, patches = ax.hist(angles, bins=45, range=(0, 90),
                                   color='#4D9EFF', edgecolor='#8AAAD0',
                                   linewidth=0.5, alpha=0.85)
        
        # Renk gradyanı (açıya göre)
        for i, patch in enumerate(patches):
            angle_val = (bins[i] + bins[i+1]) / 2
            if angle_val < 30:
                patch.set_facecolor('#2ECC71')  # Yeşil - iyi
            elif angle_val < 60:
                patch.set_facecolor('#F39C12')  # Turuncu - orta
            else:
                patch.set_facecolor('#E74C3C')  # Kırmızı - kötü
        
        # İstatistik çizgileri
        mean_angle = np.mean(angles)
        median_angle = np.median(angles)
        
        ax.axvline(mean_angle, color='#4D9EFF', linestyle='--', linewidth=2,
                  label=f'Ortalama: {mean_angle:.1f}°', zorder=10)
        ax.axvline(median_angle, color='#F39C12', linestyle=':', linewidth=2,
                  label=f'Medyan: {median_angle:.1f}°', zorder=10)
        
        # Etiketler
        ax.set_xlabel('Gelme Açısı (°)', color=text_color, fontsize=11, fontweight='bold')
        ax.set_ylabel('Üçgen Sayısı', color=text_color, fontsize=11, fontweight='bold')
        ax.set_xlim(0, 90)
        ax.set_title('0° = Dik Çarpış  →  90° = Sıyırma',
                    color='#666666', fontsize=10, pad=10)
        
        # Legend
        ax.legend(fontsize=10, facecolor='#F5F5F5', labelcolor=text_color,
                 edgecolor='#D0D0D0', loc='upper right')
        
        # Grid
        ax.grid(axis='y', linestyle='--', alpha=0.3, color='#D0D0D0')
        ax.set_axisbelow(True)
    else:
        ax.text(0.5, 0.5, 'Veri yok', transform=ax.transAxes,
               ha='center', va='center', fontsize=14, color=text_color)
    
    # Stil
    ax.tick_params(colors=text_color, labelsize=9)
    ax.spines['bottom'].set_color('#D0D0D0')
    ax.spines['left'].set_color('#D0D0D0')
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    
    fig.suptitle('Gelme Açısı Dağılımı', color=text_color, fontsize=13, fontweight='bold')
    fig.tight_layout()
    
    return fig


def create_incoming_angle_histogram(resource_data, bg_color="#FFFFFF", text_color="#333333"):
    """
    Gelen açı histogramı oluştur (tüm çarpışmalar için).
    0° = Dik çarpış, 90° = Sıyırma
    """
    fig = plt.Figure(figsize=(10, 6), facecolor=bg_color, dpi=80)
    ax = fig.add_subplot(111)
    ax.set_facecolor('#FFFFFF')
    
    if resource_data and "incoming_angles" in resource_data:
        angles = resource_data["incoming_angles"]
        
        if len(angles) > 0:
            # Histogram
            n, bins, patches = ax.hist(angles, bins=90, range=(0, 90),
                                       color='#4D9EFF', edgecolor='#8AAAD0',
                                       linewidth=0.5, alpha=0.85)
            
            # Renk gradyanı (açıya göre)
            for i, patch in enumerate(patches):
                angle_val = (bins[i] + bins[i+1]) / 2
                if angle_val < 30:
                    patch.set_facecolor('#2ECC71')  # Yeşil - iyi
                elif angle_val < 60:
                    patch.set_facecolor('#F39C12')  # Turuncu - orta
                else:
                    patch.set_facecolor('#E74C3C')  # Kırmızı - kötü
            
            # İstatistik çizgileri
            mean_angle = np.mean(angles)
            median_angle = np.median(angles)
            
            ax.axvline(mean_angle, color='#4D9EFF', linestyle='--', linewidth=2,
                      label=f'Ortalama: {mean_angle:.1f}°', zorder=10)
            ax.axvline(median_angle, color='#F39C12', linestyle=':', linewidth=2,
                      label=f'Medyan: {median_angle:.1f}°', zorder=10)
            
            # Etiketler
            ax.set_xlabel('Gelen Açı (°)', color=text_color, fontsize=11, fontweight='bold')
            ax.set_ylabel('Çarpışma Sayısı', color=text_color, fontsize=11, fontweight='bold')
            ax.set_xlim(0, 90)
            ax.set_title(f'Toplam {len(angles):,} çarpışma  |  0° = Dik  →  90° = Sıyırma',
                        color='#666666', fontsize=10, pad=10)
            
            # Legend
            ax.legend(fontsize=10, facecolor='#F5F5F5', labelcolor=text_color,
                     edgecolor='#D0D0D0', loc='upper right')
            
            # Grid
            ax.grid(axis='y', linestyle='--', alpha=0.3, color='#D0D0D0')
            ax.set_axisbelow(True)
        else:
            ax.text(0.5, 0.5, 'Veri yok', transform=ax.transAxes,
                   ha='center', va='center', fontsize=14, color=text_color)
    else:
        ax.text(0.5, 0.5, 'Veri yok', transform=ax.transAxes,
               ha='center', va='center', fontsize=14, color=text_color)
    
    # Stil
    ax.tick_params(colors=text_color, labelsize=9)
    ax.spines['bottom'].set_color('#D0D0D0')
    ax.spines['left'].set_color('#D0D0D0')
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    
    fig.suptitle('Gelen Açı Dağılımı', color=text_color, fontsize=13, fontweight='bold')
    fig.tight_layout()
    
    return fig


def create_reflection_angle_histogram(resource_data, bg_color="#FFFFFF", text_color="#333333"):
    """
    Seken açı histogramı oluştur (tüm çarpışmalar için).
    0° = Dik sekme, 90° = Sıyırma
    """
    fig = plt.Figure(figsize=(10, 6), facecolor=bg_color, dpi=80)
    ax = fig.add_subplot(111)
    ax.set_facecolor('#FFFFFF')
    
    if resource_data and "reflection_angles" in resource_data:
        angles = resource_data["reflection_angles"]
        
        if len(angles) > 0:
            # Histogram
            n, bins, patches = ax.hist(angles, bins=90, range=(0, 90),
                                       color='#9B59B6', edgecolor='#BB8FCE',
                                       linewidth=0.5, alpha=0.85)
            
            # Renk gradyanı (açıya göre)
            for i, patch in enumerate(patches):
                angle_val = (bins[i] + bins[i+1]) / 2
                if angle_val < 30:
                    patch.set_facecolor('#8E44AD')  # Mor - dik sekme
                elif angle_val < 60:
                    patch.set_facecolor('#AF7AC5')  # Açık mor - orta
                else:
                    patch.set_facecolor('#D7BDE2')  # Çok açık mor - sıyırma
            
            # İstatistik çizgileri
            mean_angle = np.mean(angles)
            median_angle = np.median(angles)
            
            ax.axvline(mean_angle, color='#8E44AD', linestyle='--', linewidth=2,
                      label=f'Ortalama: {mean_angle:.1f}°', zorder=10)
            ax.axvline(median_angle, color='#AF7AC5', linestyle=':', linewidth=2,
                      label=f'Medyan: {median_angle:.1f}°', zorder=10)
            
            # Etiketler
            ax.set_xlabel('Seken Açı (°)', color=text_color, fontsize=11, fontweight='bold')
            ax.set_ylabel('Çarpışma Sayısı', color=text_color, fontsize=11, fontweight='bold')
            ax.set_xlim(0, 90)
            ax.set_title(f'Toplam {len(angles):,} çarpışma  |  0° = Dik Sekme  →  90° = Sıyırma',
                        color='#666666', fontsize=10, pad=10)
            
            # Legend
            ax.legend(fontsize=10, facecolor='#F5F5F5', labelcolor=text_color,
                     edgecolor='#D0D0D0', loc='upper right')
            
            # Grid
            ax.grid(axis='y', linestyle='--', alpha=0.3, color='#D0D0D0')
            ax.set_axisbelow(True)
        else:
            ax.text(0.5, 0.5, 'Veri yok', transform=ax.transAxes,
                   ha='center', va='center', fontsize=14, color=text_color)
    else:
        ax.text(0.5, 0.5, 'Veri yok', transform=ax.transAxes,
               ha='center', va='center', fontsize=14, color=text_color)
    
    # Stil
    ax.tick_params(colors=text_color, labelsize=9)
    ax.spines['bottom'].set_color('#D0D0D0')
    ax.spines['left'].set_color('#D0D0D0')
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    
    fig.suptitle('Seken Açı Dağılımı', color=text_color, fontsize=13, fontweight='bold')
    fig.tight_layout()
    
    return fig


def create_bounce_histogram(bounce_paths, A, B, C, hit_counts, avg_angles,
                           bg_color="#FFFFFF", text_color="#333333"):
    """Sekme açısı histogramı oluştur.
    Her sekme için ortalama açıyı gösterir."""
    fig = plt.Figure(figsize=(10, 6), facecolor=bg_color, dpi=80)
    ax = fig.add_subplot(111)
    ax.set_facecolor('#FFFFFF')
    
    if len(bounce_paths) > 0:
        bounce_numbers = list(range(1, len(bounce_paths) + 1))
        bounce_angles = []
        bounce_counts = []
        
        # Her sekme için istatistik hesapla
        for bounce_idx, (start_pts, end_pts) in enumerate(bounce_paths):
            n_rays = start_pts.shape[0]
            bounce_counts.append(n_rays)
            # Ortalama açı (tüm ışınlar için)
            if n_rays > 0:
                avg_angle = 45.0  # Placeholder - gerçek hesaplama için daha fazla veri gerekli
                bounce_angles.append(avg_angle)
            else:
                bounce_angles.append(0)
        
        # Bar chart
        bars = ax.bar(bounce_numbers, bounce_counts,
                     color='#4D9EFF', edgecolor='#8AAAD0',
                     linewidth=1, alpha=0.85, width=0.7)
        
        # Renk gradyanı
        max_count = max(bounce_counts) if bounce_counts else 1
        for i, bar in enumerate(bars):
            intensity = bounce_counts[i] / max_count
            bar.set_facecolor(plt.cm.viridis(intensity))
        
        # Etiketler
        ax.set_xlabel('Sekme Numarası', color=text_color, fontsize=11, fontweight='bold')
        ax.set_ylabel('Işın Sayısı', color=text_color, fontsize=11, fontweight='bold')
        ax.set_title(f'Toplam {len(bounce_paths)} Sekme',
                    color='#666666', fontsize=10, pad=10)
        
        # Grid
        ax.grid(axis='y', linestyle='--', alpha=0.3, color='#D0D0D0')
        ax.set_axisbelow(True)
        
        # X ekseni tam sayı
        ax.set_xticks(bounce_numbers)
    else:
        ax.text(0.5, 0.5, 'Veri yok', transform=ax.transAxes,
               ha='center', va='center', fontsize=14, color=text_color)
    
    # Stil
    ax.tick_params(colors=text_color, labelsize=9)
    ax.spines['bottom'].set_color('#D0D0D0')
    ax.spines['left'].set_color('#D0D0D0')
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    
    fig.suptitle('Sekme Başına Işın Dağılımı', color=text_color, fontsize=13, fontweight='bold')
    fig.tight_layout()
    
    return fig


def create_cpu_graph(resource_data, bg_color="#FFFFFF", text_color="#333333"):
    """CPU kullanım grafiği oluştur."""
    fig = plt.Figure(figsize=(10, 6), facecolor=bg_color, dpi=80)
    ax = fig.add_subplot(111)
    ax.set_facecolor('#FFFFFF')
    
    if resource_data and len(resource_data.get("cpu", [])) > 1:
        cpu_arr = np.array(resource_data["cpu"])
        t_arr = np.array(resource_data["times"])
        
        # Ana çizgi
        ax.plot(t_arr, cpu_arr, color='#4D9EFF', linewidth=2, alpha=0.9, label='CPU Kullanımı')
        ax.fill_between(t_arr, cpu_arr, alpha=0.2, color='#4D9EFF')
        
        # İstatistik çizgileri
        mean_cpu = np.mean(cpu_arr)
        max_cpu = np.max(cpu_arr)
        
        ax.axhline(mean_cpu, color='#2ECC71', linestyle='--', linewidth=1.5,
                  label=f'Ortalama: {mean_cpu:.1f}%', zorder=10)
        ax.axhline(max_cpu, color='#E74C3C', linestyle=':', linewidth=1.5,
                  label=f'Maksimum: {max_cpu:.1f}%', zorder=10)
        
        # Etiketler
        ax.set_xlabel('Süre (saniye)', color=text_color, fontsize=11, fontweight='bold')
        ax.set_ylabel('CPU Kullanımı (%)', color=text_color, fontsize=11, fontweight='bold')
        ax.set_xlim(0, t_arr[-1])
        ax.set_ylim(0, max(100, max_cpu * 1.1))
        
        # Legend
        ax.legend(fontsize=10, facecolor='#F5F5F5', labelcolor=text_color,
                 edgecolor='#D0D0D0', loc='upper right')
        
        # Grid
        ax.grid(linestyle='--', alpha=0.3, color='#D0D0D0')
        ax.set_axisbelow(True)
    else:
        ax.text(0.5, 0.5, 'Veri yok', transform=ax.transAxes,
               ha='center', va='center', fontsize=14, color=text_color)
    
    # Stil
    ax.tick_params(colors=text_color, labelsize=9)
    ax.spines['bottom'].set_color('#D0D0D0')
    ax.spines['left'].set_color('#D0D0D0')
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    
    fig.suptitle('CPU Kullanımı — Zaman Serisi', color=text_color, fontsize=13, fontweight='bold')
    fig.tight_layout()
    
    return fig


def create_ram_graph(resource_data, bg_color="#FFFFFF", text_color="#333333"):
    """
    RAM kullanım grafiği oluştur.
    PERFORMANS OPTİMİZE EDİLDİ: Daha düşük DPI.
    """
    fig = plt.Figure(figsize=(10, 6), facecolor=bg_color, dpi=80)
    ax = fig.add_subplot(111)
    ax.set_facecolor('#FFFFFF')
    
    if resource_data and len(resource_data.get("ram", [])) > 1:
        ram_arr = np.array(resource_data["ram"]) / 1024.0  # MB -> GB
        t_arr = np.array(resource_data["times"])
        
        # Ana çizgi
        ax.plot(t_arr, ram_arr, color='#2ECC71', linewidth=2, alpha=0.9, label='RAM Kullanımı')
        ax.fill_between(t_arr, ram_arr, alpha=0.2, color='#2ECC71')
        
        # İstatistik çizgileri
        mean_ram = np.mean(ram_arr)
        max_ram = np.max(ram_arr)
        
        ax.axhline(mean_ram, color='#4A90E2', linestyle='--', linewidth=1.5,
                  label=f'Ortalama: {mean_ram:.2f} GB', zorder=10)
        ax.axhline(max_ram, color='#E74C3C', linestyle=':', linewidth=1.5,
                  label=f'Maksimum: {max_ram:.2f} GB', zorder=10)
        
        # Etiketler
        ax.set_xlabel('Süre (saniye)', color=text_color, fontsize=11, fontweight='bold')
        ax.set_ylabel('RAM Kullanımı (GB)', color=text_color, fontsize=11, fontweight='bold')
        ax.set_xlim(0, t_arr[-1])
        ax.set_ylim(0, max_ram * 1.1)
        
        # Legend
        ax.legend(fontsize=10, facecolor='#F5F5F5', labelcolor=text_color,
                 edgecolor='#D0D0D0', loc='upper right')
        
        # Grid
        ax.grid(linestyle='--', alpha=0.3, color='#D0D0D0')
        ax.set_axisbelow(True)
    else:
        ax.text(0.5, 0.5, 'Veri yok', transform=ax.transAxes,
               ha='center', va='center', fontsize=14, color=text_color)
    
    # Stil
    ax.tick_params(colors=text_color, labelsize=9)
    ax.spines['bottom'].set_color('#D0D0D0')
    ax.spines['left'].set_color('#D0D0D0')
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    
    fig.suptitle('RAM Kullanımı — Zaman Serisi', color=text_color, fontsize=13, fontweight='bold')
    fig.tight_layout()
    
    return fig


# ------------------------------------------------
# 9. EXCEL KAYIT FONKSİYONU
# ------------------------------------------------
def save_detailed_results_to_excel(hit_counts, avg_angles, n_theta, n_phi,
                                    filename=None, resource_data=None):
    """
    Simülasyon sonuçlarını mükerrer veri üretmeden optimize bir şekilde kaydeder.
    Sayfa 1: Simülasyon_Raporu -> Sadece İstatistiksel Özet (Tablosuz, Temiz Grafik/Metrik Alanı)
    Sayfa 2: Vurus_Alan_Ucgenler -> Sadece hit_count > 0 olan veri detayları (Azalan Sırada)
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("\n[UYARI] openpyxl kütüphanesi bulunamadı: pip install openpyxl")
        return
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if filename is None:
        filename = f"Isin_Analiz_Raporu_{timestamp}.xlsx"

    print(f"\n[EXCEL] Rapor optimize edilerek kaydediliyor: {filename}")

    active_hits = hit_counts > 0
    active_indices = np.where(active_hits)[0]
    total_sent_rays = n_theta * n_phi
    total_bounces = int(np.sum(hit_counts))

    try:
        import pandas as pd
        has_pandas = True
    except ImportError:
        has_pandas = False

    wb = openpyxl.Workbook()


    # EXCEL SAYFA 1

    ws1 = wb.active
    ws1.title = "Simülasyon_Raporu"
    ws1.views.sheetView[0].showGridLines = True

    # Stil Tanımlamaları
    title_font = Font(name="Segoe UI", size=15, bold=True, color="1F4E78")
    section_font = Font(name="Segoe UI", size=11, bold=True, color="2B579A")
    label_font = Font(name="Segoe UI", size=10, bold=False, color="555555")
    value_font = Font(name="Segoe UI", size=10, bold=True, color="111111")
    
    border_thin = Border(bottom=Side(style='thin', color='E0E0E0'))

    # Başlık Alanı
    ws1['A1'] = "IŞIN SİMÜLASYONU ÖZET RAPORU"
    ws1['A1'].font = title_font
    ws1.row_dimensions[1].height = 25

    metrics = [
        ("Genel Bilgiler", ""),
        ("Tarih / Saat", datetime.now().strftime('%d.%m.%Y %H:%M:%S')),
        ("Toplam Gönderilen Işın", total_sent_rays),
        ("Sistemdeki Toplam Üçgen", len(hit_counts)),
        
        ("Sekme & Etkileşim İstatistikleri", ""),
        ("Işın İsabet Eden Benzersiz Üçgen", len(active_indices)),
        ("Toplam Sekme (Vuruş) Sayısı", total_bounces),
        
        ("Açısal Analiz Özeti (Gelme/Sekme)", ""),
    ]

    if len(active_indices) > 0:
        angles_only = avg_angles[active_indices]
        metrics.extend([
            ("Ortalama Açı (°)", round(float(np.mean(angles_only)), 2)),
            ("Minimum Açı (°)", round(float(np.min(angles_only)), 2)),
            ("Maksimum Açı (°)", round(float(np.max(angles_only)), 2)),
            ("Standart Sapma (°)", round(float(np.std(angles_only)), 2)),
            ("Açı Varyansı", round(float(np.var(angles_only)), 2)),
            ("En Çok Darbe Alan Üçgen ID", int(active_indices[np.argmax(hit_counts[active_indices])] + 1)),
            ("En Yüksek Vuruş Sayısı", int(np.max(hit_counts[active_indices])))
        ])

    row_idx = 3
    for label, val in metrics:
        if val == "":  # Kategori Başlığı
            row_idx += 1
            cell = ws1.cell(row=row_idx, column=1, value=label)
            cell.font = section_font
            ws1.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
            row_idx += 1
        else:          # Veri Satırı
            c_label = ws1.cell(row=row_idx, column=1, value=label)
            c_label.font = label_font
            c_label.border = border_thin
            
            c_val = ws1.cell(row=row_idx, column=2, value=val)
            c_val.font = value_font
            c_val.alignment = Alignment(horizontal="right")
            c_val.border = border_thin
            row_idx += 1

    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 25


    # EXCEL SAYFA 2
    
    if len(active_indices) > 0:
        ws2 = wb.create_sheet("Vurus_Alan_Ucgenler")
        
        hit_data = {
            "Üçgen ID": (active_indices + 1).tolist(),  # 1-based index
            "Vuruş Sayısı": hit_counts[active_indices].tolist(),
            "Ort. Gelme/Sekme Açısı (°)": [round(float(x), 4) for x in avg_angles[active_indices]],
        }
        
        # Pandas ile dataframe oluşturup vuruş sayısına göre azalan sırala
        df_hits = pd.DataFrame(hit_data).sort_values(by="Vuruş Sayısı", ascending=False)
        
        # Tablo Tasarım Stilleri
        header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        data_font = Font(name="Segoe UI", size=10)
        align_center = Alignment(horizontal="center", vertical="center")
        border_data = Border(
            left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'), bottom=Side(style='thin', color='E0E0E0')
        )

        # Başlıkları Yazdır
        for c_idx, col_name in enumerate(df_hits.columns, start=1):
            cell = ws2.cell(row=1, column=c_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center

        # Verileri Yazdır
        for r_idx, row_values in enumerate(df_hits.values, start=2):
            # Durumsal Renklendirme fill mantığı (Açıya göre soft renk geçişi)
            angle_val = row_values[2]
            if angle_val < 30:
                color_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft Yeşil
            elif angle_val < 60:
                color_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Soft Sarı
            else:
                color_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft Kırmızı

            for c_idx, val in enumerate(row_values, start=1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=val)
                cell.font = data_font
                cell.alignment = align_center
                cell.border = border_data
                if c_idx == 3:  # Sadece açı hücresini duruma göre renklendir (Tüm satırı boyayıp göz yormasın)
                    cell.fill = color_fill

        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 25

    # Kaydet
    try:
        wb.save(filename)
        print(f"[EXCEL] ✓ Temiz rapor başarıyla oluşturuldu: {filename}")
        return filename
    except Exception as e:
        print(f"[EXCEL] ✗ Kayıt hatası: {e}")
        return None
